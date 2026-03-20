"""Tests for DFFH rental Excel extraction (synthetic xlsx)."""
import os
import tempfile

import openpyxl

from ingestion.dffh_rental import extract_latest_rents


def _build_minimal_dffh_xlsx(path: str) -> None:
    """
    Mirrors current DFFH layout:
    - Sheet names: '3 bedroom house', '2 bedroom flat' (or spelled-out variants — tested elsewhere)
    - Row 1: quarter labels (pairs: Count label, Median label — same text repeated)
    - Row 2: Count / Median / Count / Median …
    - Data from row 3+: col 1 = compound suburb, latest Median column = rent
    """
    wb = openpyxl.Workbook()
    default = wb.active
    wb.remove(default)

    for name in ("3 bedroom house", "2 bedroom flat"):
        ws = wb.create_sheet(title=name)
        ws.append(["title", None, None, None, None, None])
        ws.append([None, None, "Jun 2025", "Jun 2025", "Sep 2025", "Sep 2025"])
        ws.append([None, None, "Count", "Median", "Count", "Median"])
        ws.append(["", "Albert Park-Middle Park", 1, 500, 1, 600])
        ws.append(["", "Solo Suburb", 1, 520, 1, 550])

    wb.save(path)


def test_extract_latest_rents_splits_compound_and_both_sheets():
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
        path = tmp.name
    try:
        _build_minimal_dffh_xlsx(path)
        out = extract_latest_rents(path)
    finally:
        os.unlink(path)

    assert "ALBERT PARK" in out
    assert "MIDDLE PARK" in out
    assert out["ALBERT PARK"]["rent_3br_wk"] == 600
    assert out["ALBERT PARK"]["rent_2br_wk"] == 600
    assert out["SOLO SUBURB"]["rent_3br_wk"] == 550
    assert out["SOLO SUBURB"]["rent_2br_wk"] == 550


def test_extract_latest_rents_finds_latest_2025_median_column():
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
        path = tmp.name
    try:
        wb = openpyxl.Workbook()
        wb.remove(wb.active)
        ws = wb.create_sheet("3 bedroom house")
        ws.append(["a", "b", "c", "d", "e", "f"])
        ws.append(["", "", "Jun 2025", "Jun 2025", "Sep 2025", "Sep 2025"])
        ws.append(["", "", "Count", "Median", "Count", "Median"])
        ws.append(["", "Testville", 1, 500, 1, 720])
        wb.save(path)
        out = extract_latest_rents(path)
    finally:
        os.unlink(path)

    assert out["TESTVILLE"]["rent_3br_wk"] == 720


def test_extract_resolves_three_bedroom_sheet_name():
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
        path = tmp.name
    try:
        wb = openpyxl.Workbook()
        wb.remove(wb.active)
        ws = wb.create_sheet("Three bedroom house")
        ws.append(["a", "b", "c", "d", "e", "f"])
        ws.append(["", "", "Sep 2025", "Sep 2025", None, None])
        ws.append(["", "", "Count", "Median", None, None])
        ws.append(["", "Spelled Suburb", 1, 800, None, None])
        wb.save(path)
        out = extract_latest_rents(path)
    finally:
        os.unlink(path)

    assert out["SPELLED SUBURB"]["rent_3br_wk"] == 800


def test_extract_ignores_low_rent_noise():
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
        path = tmp.name
    try:
        wb = openpyxl.Workbook()
        wb.remove(wb.active)
        ws = wb.create_sheet("3 bedroom house")
        ws.append(["a", "b", "c", "d"])
        ws.append(["", "", "Sep 2025", "Sep 2025"])
        ws.append(["", "", "Count", "Median"])
        ws.append(["", "Low Rent Suburb", 1, 30])  # < 50 skipped
        wb.save(path)
        out = extract_latest_rents(path)
    finally:
        os.unlink(path)

    assert "LOW RENT SUBURB" not in out
