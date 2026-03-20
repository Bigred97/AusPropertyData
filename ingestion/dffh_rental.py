"""
DFFH rental report ingestion.
Downloads latest Excel from dffh.vic.gov.au, extracts Sep-quarter medians,
splits compound suburb names, upserts rent columns into suburbs table.
"""
import asyncio
import asyncpg
import openpyxl
import os
from bs4 import BeautifulSoup

from ingestion.fetch import first_ckan_go_to_resource, ingestion_http_client

CATALOGUE_URL = "https://discover.data.vic.gov.au/dataset/rental-report-quarterly-moving-annual-rents-by-suburb"


def _require_db_url() -> str:
    url = os.environ.get("SUPABASE_DB_URL")
    if not url:
        raise RuntimeError("SUPABASE_DB_URL environment variable is required for DFFH DB ingestion")
    return url


async def get_rental_download_url() -> str:
    async with ingestion_http_client() as client:
        resp = await client.get(CATALOGUE_URL, timeout=30)
        resp.raise_for_status()
        soup = BeautifulSoup(resp.text, "html.parser")
    hit = first_ckan_go_to_resource(
        soup,
        CATALOGUE_URL,
        href_predicate=lambda h: "dffh.vic.gov.au" in h or "moving" in h.lower(),
    )
    if not hit:
        raise ValueError("Could not find DFFH rental download link")
    return hit[0]


def _resolve_sheet_name(wb, *candidates: str) -> str | None:
    """Match DFFH sheet tabs (names vary e.g. '3 bedroom house' vs 'Three bedroom house')."""
    by_lower = {n.lower(): n for n in wb.sheetnames}
    for c in candidates:
        if c in wb.sheetnames:
            return c
        if c.lower() in by_lower:
            return by_lower[c.lower()]
    return None


def extract_latest_rents(filepath: str) -> dict[str, dict]:
    """
    Returns {suburb_upper: {rent_3br_wk: X, rent_2br_wk: Y}}
    Splits compound suburb names e.g. 'Albert Park-Middle Park' → both suburbs get same rent.
    """
    wb = openpyxl.load_workbook(filepath, read_only=True)
    result = {}

    sheet_map = [
        (["3 bedroom house", "three bedroom house"], "rent_3br_wk"),
        (["2 bedroom flat", "two bedroom flat"], "rent_2br_wk"),
    ]

    for candidates, field_name in sheet_map:
        sheet_name = _resolve_sheet_name(wb, *candidates)
        if not sheet_name:
            continue
        ws = wb[sheet_name]
        all_rows = list(ws.iter_rows(values_only=True))

        # Latest year's median column: row 1 = quarter labels, row 2 = Count/Median
        header_row = all_rows[1]
        type_row = all_rows[2] if len(all_rows) > 2 else ()
        latest_col = None
        for i in range(len(header_row) - 1, -1, -1):
            lab = header_row[i]
            typ = type_row[i] if i < len(type_row) else None
            if lab and typ == "Median" and "2025" in str(lab):
                latest_col = i
                break
        if latest_col is None:
            continue

        for row in all_rows[3:]:
            compound = row[1]
            rent = row[latest_col] if latest_col < len(row) else None
            if not compound or not isinstance(rent, (int, float)) or rent < 50:
                continue
            # Split compound names: "Albert Park-Middle Park-West St Kilda"
            parts = [
                p.strip().upper()
                for p in str(compound).split("-")
                if len(p.strip()) > 2
            ]
            for suburb in parts:
                if suburb not in result:
                    result[suburb] = {}
                result[suburb][field_name] = int(rent)

    wb.close()
    return result


async def compute_yield_update(conn, suburb: str, rent_3br: int) -> None:
    """Update gross_yield based on latest rent and price."""
    row = await conn.fetchrow(
        "SELECT price_2023 FROM suburbs WHERE suburb = $1", suburb
    )
    if row and row["price_2023"] and rent_3br:
        yield_ = round((rent_3br * 52 / row["price_2023"]) * 100, 2)
        await conn.execute(
            """
            UPDATE suburbs SET rent_3br_wk=$2, gross_yield=$3, data_updated_at=NOW()
            WHERE suburb=$1
        """,
            suburb,
            rent_3br,
            yield_,
        )


async def ingest_dffh_rental():
    print("DFFH: loading rental data...")
    local = os.environ.get("DFFH_RENTAL_XLSX")
    mirror = (os.environ.get("DFFH_RENTAL_URL") or "").strip()
    if local and os.path.isfile(local):
        filepath = os.path.abspath(local)
        print(f"  Using local file: {filepath}")
    elif mirror:
        print(f"  Using DFFH_RENTAL_URL mirror: {mirror[:100]}…")
        async with ingestion_http_client() as client:
            resp = await client.get(mirror, timeout=120)
            resp.raise_for_status()
        filepath = "/tmp/dffh_rental.xlsx"
        with open(filepath, "wb") as f:
            f.write(resp.content)
    else:
        download_url = await get_rental_download_url()
        print(f"  Download: {download_url}")
        async with ingestion_http_client() as client:
            resp = await client.get(download_url, timeout=300)
            resp.raise_for_status()
        filepath = "/tmp/dffh_rental.xlsx"
        with open(filepath, "wb") as f:
            f.write(resp.content)

    rents = extract_latest_rents(filepath)
    print(f"  Extracted rents for {len(rents)} suburb entries")

    conn = await asyncpg.connect(_require_db_url(), statement_cache_size=0)
    updated = 0
    for suburb, fields in rents.items():
        rent_3br = fields.get("rent_3br_wk")
        rent_2br = fields.get("rent_2br_wk")
        if rent_3br:
            await compute_yield_update(conn, suburb, rent_3br)
            updated += 1
        if rent_2br:
            await conn.execute(
                "UPDATE suburbs SET rent_2br_wk=$2, data_updated_at=NOW() WHERE suburb=$1",
                suburb,
                rent_2br,
            )
    await conn.execute(
        "INSERT INTO ingestion_log (source, status, rows_upserted) VALUES ('dffh_rental', 'success', $1)",
        updated,
    )
    await conn.close()
    print(f"  Updated {updated} suburbs with rental data")


if __name__ == "__main__":
    from ingestion.project_env import load_project_dotenv

    load_project_dotenv()
    asyncio.run(ingest_dffh_rental())
